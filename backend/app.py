# app.py
import os
import io
import uuid
import fitz
import secrets
import hashlib
import datetime as dt
from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS
from flask_jwt_extended import JWTManager, create_access_token, jwt_required, get_jwt
from pymongo import MongoClient, ASCENDING, errors
import bcrypt
from fpdf import FPDF
from tempfile import NamedTemporaryFile
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from google.cloud import vision
import docx2txt
from extract_fields_from_pdf import extract_fields_from_pdf
from dotenv import load_dotenv

# ---------------------------
# Load .env
# ---------------------------
load_dotenv()
MONGODB_URI = os.getenv("MONGODB_URI", "mongodb://localhost:27017")
DB_NAME = os.getenv("DB_NAME", "pdf_ocr_db")
JWT_SECRET = os.getenv("JWT_SECRET_KEY", "change-this-in-production")
GCP_CREDS = os.getenv("GOOGLE_APPLICATION_CREDENTIALS", None)
if GCP_CREDS:
    os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = GCP_CREDS

# ---------------------------
# Flask + JWT + CORS
# ---------------------------
app = Flask(__name__, static_folder="build", static_url_path="/")
CORS(app)
app.config["JWT_SECRET_KEY"] = JWT_SECRET
jwt = JWTManager(app)

# ---------------------------
# MongoDB
# ---------------------------
client = MongoClient(MONGODB_URI)
db = client[DB_NAME]
users_coll = db["users"]
invites_coll = db["invites"]

try:
    users_coll.create_index([("username", ASCENDING)], unique=True)
    invites_coll.create_index([("token_hash", ASCENDING)], unique=True)
    invites_coll.create_index("expires_at", expireAfterSeconds=0)
except Exception:
    pass

# ---------------------------
# RBAC helpers
# ---------------------------
ROLES = {"Admin", "User", "Viewer"}


def hash_password(plain: str) -> bytes:
    # returns bcrypt hash bytes
    return bcrypt.hashpw(plain.encode("utf-8"), bcrypt.gensalt())


def check_password(plain: str, stored) -> bool:
    # stored may be bytes or str
    if stored is None:
        return False
    stored_bytes = stored.encode("utf-8") if isinstance(stored, str) else stored
    return bcrypt.checkpw(plain.encode("utf-8"), stored_bytes)


def role_required(allowed_roles):
    def outer(fn):
        from functools import wraps

        @wraps(fn)
        @jwt_required()
        def wrapper(*args, **kwargs):
            claims = get_jwt() or {}
            if claims.get("role") not in allowed_roles:
                return jsonify({"error": "Access forbidden: Insufficient role"}), 403
            return fn(*args, **kwargs)

        return wrapper

    return outer


# ---------------------------
# Invite helpers
# ---------------------------
INVITE_TTL_HOURS = int(os.getenv("INVITE_TTL_HOURS", "48"))
ALLOWED_INVITE_ROLES = {"Admin", "Viewer"}


def hash_token(token: str) -> str:
    return hashlib.sha256(token.encode("utf-8")).hexdigest()


def create_invite(username: str, role: str) -> str:
    if role not in ALLOWED_INVITE_ROLES:
        raise ValueError("Invalid role for invite")
    username = username.strip()
    for _ in range(5):
        token = secrets.token_urlsafe(32)
        token_hash = hash_token(token)
        now = dt.datetime.utcnow()
        expires_at = now + dt.timedelta(hours=INVITE_TTL_HOURS)
        try:
            invites_coll.insert_one({
                "username": username,
                "role": role,
                "token_hash": token_hash,
                "created_at": now,
                "expires_at": expires_at,
                "used_at": None
            })
            return token
        except errors.DuplicateKeyError:
            continue
    raise RuntimeError("Failed to generate unique invite token")


def consume_invite(username: str, raw_token: str):
    if not raw_token:
        return None
    token_hash = hash_token(raw_token)
    username = (username or "").strip()
    invite = invites_coll.find_one({"token_hash": token_hash, "username": username})
    if not invite or invite.get("used_at"):
        return None
    if invite.get("expires_at") and invite["expires_at"] < dt.datetime.utcnow():
        return None
    invites_coll.update_one({"_id": invite["_id"]}, {"$set": {"used_at": dt.datetime.utcnow()}})
    return invite.get("role")


def peek_invite(username: str, raw_token: str):
    if not raw_token:
        return None
    token_hash = hash_token(raw_token)
    username = (username or "").strip()
    invite = invites_coll.find_one({
        "token_hash": token_hash,
        "username": username,
        "used_at": None
    })
    if not invite:
        return None
    if invite.get("expires_at") and invite["expires_at"] < dt.datetime.utcnow():
        return None
    return invite.get("role")


# ---------------------------
# Auth endpoints
# ---------------------------
@app.post("/auth/signup")
def auth_signup():
    data = request.get_json(silent=True) or {}
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""
    invite_code = (data.get("invite_code") or "").strip()
    if not username or not password:
        return jsonify({"error": "Username and password required"}), 400
    try:
        # First user becomes Admin by default
        if users_coll.count_documents({}) == 0:
            role = "Admin"
        else:
            role = "User"

        if invite_code:
            invited_role = consume_invite(username, invite_code)
            if invited_role:
                role = invited_role
            else:
                return jsonify({"error": "Invalid or expired invite code"}), 400

        hashed = hash_password(password)
        users_coll.insert_one({"username": username, "password": hashed, "role": role})
        return jsonify({"message": "User created", "username": username, "role": role}), 201
    except errors.DuplicateKeyError:
        return jsonify({"error": "User exists"}), 409
    except Exception as e:
        return jsonify({"error": "Signup failed", "details": str(e)}), 500


@app.post("/auth/login")
def auth_login():
    data = request.get_json(silent=True) or {}
    username, password = (data.get("username") or "").strip(), data.get("password") or ""
    if not username or not password:
        return jsonify({"error": "Username and password required"}), 400
    user = users_coll.find_one({"username": username})
    if not user or not check_password(password, user.get("password")):
        return jsonify({"error": "Invalid credentials"}), 401
    role = user.get("role", "User")
    token = create_access_token(identity=username, additional_claims={"role": role, "sub": username})
    return jsonify({"token": token, "role": role, "username": username})


@app.get("/auth/me")
@jwt_required()
def auth_me():
    claims = get_jwt()
    return jsonify({"username": claims.get("sub"), "role": claims.get("role")})


@app.post("/auth/admin/invite")
@role_required(["Admin"])
def admin_invite():
    data = request.get_json(silent=True) or {}
    username = (data.get("username") or "").strip()
    role = (data.get("role") or "Admin").strip()
    if not username:
        return jsonify({"error": "Username is required"}), 400
    if role not in ALLOWED_INVITE_ROLES:
        return jsonify({"error": f"Role must be one of {sorted(list(ALLOWED_INVITE_ROLES))}"}), 400
    if users_coll.find_one({"username": username}):
        return jsonify({"error": "User already exists"}), 409
    try:
        token = create_invite(username, role)
        return jsonify({
            "message": f"{role} invite created",
            "username": username,
            "role": role,
            "invite_token": token,
            "expires_in_hours": INVITE_TTL_HOURS
        }), 201
    except Exception as e:
        return jsonify({"error": f"Failed to create invite: {str(e)}"}), 500


@app.post("/auth/invite/validate")
def validate_invite():
    data = request.get_json(silent=True) or {}
    username = (data.get("username") or "").strip()
    invite_code = (data.get("invite_code") or "").strip()
    if not username or not invite_code:
        return jsonify({"valid": False, "error": "username and invite_code required"}), 400
    role = peek_invite(username, invite_code)
    if role:
        return jsonify({"valid": True, "role": role}), 200
    return jsonify({"valid": False}), 200


@app.post("/signup")
def legacy_signup():
    return auth_signup()


@app.post("/login")
def legacy_login():
    return auth_login()


# ---------------------------
# OCR endpoint
# ---------------------------
@app.route("/api/ocr", methods=["POST"])
@role_required(["Admin", "User"])
def ocr_file():
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "No file uploaded"}), 400

    filename = file.filename or "uploaded.pdf"
    ext = os.path.splitext(filename.lower())[1]
    file_bytes = file.read()
    file_stream = io.BytesIO(file_bytes)

    try:
        if ext == ".pdf":
            try:
                file_stream.seek(0)
            except Exception:
                pass

            extracted = extract_fields_from_pdf(file_stream) or {}

            # canonical keys from extractor
            patients = extracted.get("patients", []) or []
            procedure_tables = extracted.get("procedure_tables", {}) or {}
            diagnosis_tables = extracted.get("diagnosis_tables", {}) or {}

            # produce legacy-friendly flattened arrays
            medical_billing_codes = []
            for section, rows in procedure_tables.items():
                for r in rows:
                    if isinstance(r, (list, tuple)):
                        code = r[0] if len(r) > 0 else ""
                        desc = r[1] if len(r) > 1 else ""
                        fee = r[2] if len(r) > 2 else ""
                    elif isinstance(r, dict):
                        code = r.get("Code") or r.get("code") or ""
                        desc = r.get("Description") or r.get("description") or ""
                        fee = r.get("Fee") or r.get("fee") or ""
                    else:
                        continue
                    medical_billing_codes.append({"Section": section, "Code": code, "Description": desc, "Fee": fee})

            diagnosis_codes = []
            for section, rows in diagnosis_tables.items():
                for r in rows:
                    if isinstance(r, (list, tuple)):
                        typ = r[0] if len(r) > 0 else "ICD-10"
                        code = r[1] if len(r) > 1 else ""
                        desc = r[2] if len(r) > 2 else ""
                    elif isinstance(r, dict):
                        typ = r.get("Type") or r.get("type") or "ICD-10"
                        code = r.get("Code") or r.get("code") or ""
                        desc = r.get("Diagnosis") or r.get("diagnosis") or r.get("Description") or r.get("description") or ""
                    else:
                        continue
                    diagnosis_codes.append({"Section": section, "Type": typ, "Code": code, "Diagnosis": desc})

            # Ensure the legacy per-patient table fields exist (extractor already attempts this)
            # but double-check: if extractor didn't assign, try to map top-level rows into patients evenly/fallback
            # (extractor should already have mapped rows to the correct patient segments)

            response = {
                "file_name": filename,
                "patients": patients,
                "procedure_tables": procedure_tables,
                "diagnosis_tables": diagnosis_tables,
                "medical_billing_codes": medical_billing_codes,
                "diagnosis_codes": diagnosis_codes,
                "metadata": extracted.get("metadata", {})
            }
            return jsonify(response)

        elif ext in [".jpg", ".jpeg", ".png"]:
            image = vision.Image(content=file_bytes)
            response = vision_client.document_text_detection(image=image) if vision_client else None
            text = (response.full_text_annotation.text if response and getattr(response, "full_text_annotation", None) else "") or ""
            flattened = {"File Name": filename, "Extracted Text": text.strip()}
            return jsonify({"file_name": filename, "patients": [flattened], "procedure_tables": {}, "diagnosis_tables": {}, "medical_billing_codes": [], "diagnosis_codes": []})

        elif ext in [".doc", ".docx"]:
            with NamedTemporaryFile(delete=False, suffix=ext) as tmp:
                tmp.write(file_bytes)
                tmp.flush()
                text = docx2txt.process(tmp.name)
            flattened = {"File Name": filename, "Extracted Text": text.strip()}
            return jsonify({"file_name": filename, "patients": [flattened], "procedure_tables": {}, "diagnosis_tables": {}, "medical_billing_codes": [], "diagnosis_codes": []})

        else:
            return jsonify({"error": f"Unsupported file type: {ext}"}), 400

    except Exception as e:
        print("[ERROR] OCR failed:", e)
        return jsonify({"error": "OCR processing failed", "details": str(e)}), 500


# Debug endpoint to inspect raw PyMuPDF text (handy to tune regexes)
@app.route("/api/ocr-debug-text", methods=["POST"])
@role_required(["Admin", "User"])
def ocr_debug_text():
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "No file uploaded"}), 400
    file_bytes = file.read()
    try:
        doc = fitz.open(stream=file_bytes, filetype="pdf")
        texts = []
        for i, page in enumerate(doc):
            txt = ""
            try:
                txt = page.get_text("text") or ""
            except Exception:
                txt = ""
            texts.append(f"--- PAGE {i+1} ---\n{txt}")
        return jsonify({"raw_text": "\n\n".join(texts)})
    except Exception as e:
        return jsonify({"error": "Failed to read PDF", "details": str(e)}), 500


# ---------------------------
# Excel export (3 sheets)
# ---------------------------
@app.route("/api/export-excel", methods=["POST", "GET"])
@role_required(["Admin", "User"])
def export_excel():
    global all_metadata
    if request.method == "POST":
        content = request.json
        if not content:
            return jsonify({"error": "No metadata to export"}), 400
        meta = content.get("metadata", content)
        if isinstance(meta, list):
            all_metadata.extend(meta)
        else:
            all_metadata.append(meta)
        return jsonify({"message": "Metadata added"})

    elif request.method == "GET":
        if not all_metadata:
            return jsonify({"error": "No metadata to export"}), 400

        wb = Workbook()

        # Sheet 1: Patients
        ws1 = wb.active
        ws1.title = "Patients"
        patient_headers = ["Patient Name", "Date of Birth", "Subscriber ID", "Primary Insurance", "Date of Service", "Patient Signature", "Physician Signature"]
        header_font = Font(bold=True, size=12)
        for i, h in enumerate(patient_headers, 1):
            ws1.cell(row=1, column=i, value=h).font = header_font
            ws1.column_dimensions[get_column_letter(i)].width = 30

        row_idx = 2
        for meta in all_metadata:
            patients = meta.get("patients") if isinstance(meta, dict) else []
            if not patients and isinstance(meta, dict) and all(k in meta for k in ["Patient Name", "Date of Birth"]):
                patients = [meta]
            for p in patients:
                for i, h in enumerate(patient_headers, 1):
                    val = p.get(h, "NIL") if isinstance(p, dict) else "NIL"
                    ws1.cell(row=row_idx, column=i, value=val)
                row_idx += 1

        # Sheet 2: Medical Billing Codes
        ws2 = wb.create_sheet(title="Billing Codes")
        billing_headers = ["Section", "Code", "Description", "Fee"]
        for i, h in enumerate(billing_headers, 1):
            ws2.cell(row=1, column=i, value=h).font = header_font
            ws2.column_dimensions[get_column_letter(i)].width = 30
        row_idx = 2
        for meta in all_metadata:
            billing = meta.get("medical_billing_codes", []) if isinstance(meta, dict) else []
            for code in billing:
                ws2.cell(row=row_idx, column=1, value=code.get("Section", ""))
                ws2.cell(row=row_idx, column=2, value=code.get("Code", ""))
                ws2.cell(row=row_idx, column=3, value=code.get("Description", ""))
                ws2.cell(row=row_idx, column=4, value=code.get("Fee", ""))
                row_idx += 1

        # Sheet 3: Diagnosis Codes
        ws3 = wb.create_sheet(title="Diagnosis Codes")
        diag_headers = ["Section", "Type", "Code", "Diagnosis"]
        for i, h in enumerate(diag_headers, 1):
            ws3.cell(row=1, column=i, value=h).font = header_font
            ws3.column_dimensions[get_column_letter(i)].width = 30
        row_idx = 2
        for meta in all_metadata:
            diag = meta.get("diagnosis_codes", []) if isinstance(meta, dict) else []
            for d in diag:
                ws3.cell(row=row_idx, column=1, value=d.get("Section", ""))
                ws3.cell(row=row_idx, column=2, value=d.get("Type", ""))
                ws3.cell(row=row_idx, column=3, value=d.get("Code", ""))
                ws3.cell(row=row_idx, column=4, value=d.get("Diagnosis", ""))
                row_idx += 1

        with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            return send_file(tmp.name, as_attachment=True, download_name="ocr_output.xlsx")


# ---------------------------
# PDF export, reset, serve React, run
# ---------------------------
class PDFReport(FPDF):
    def header(self):
        self.set_font("Arial", "B", 14)
        self.cell(0, 10, "Patient Metadata Extract", ln=True, align="C")
        self.ln(6)

    def add_metadata_table(self, metadata):
        self.set_font("Arial", "B", 12)
        self.cell(90, 10, "Field", 1)
        self.cell(100, 10, "Value", 1)
        self.ln()
        self.set_font("Arial", "", 11)
        for key, value in metadata.items():
            if key not in ["Patient Signature", "Physician Signature"]:
                val = str(value).replace("Address:", "").replace("Employer:", "").strip()
                x = self.get_x()
                y = self.get_y()
                self.multi_cell(90, 10, key, 1)
                self.set_xy(x + 90, y)
                self.multi_cell(100, 10, val, 1)
                self.ln(0)


@app.route("/api/export-pdf", methods=["POST"])
@role_required(["Admin", "User", "Viewer"])
def export_pdf():
    content = request.json
    if not content or "metadata" not in content:
        return jsonify({"error": "Missing metadata"}), 400
    pdf = PDFReport()
    pdf.add_page()
    pdf.add_metadata_table(content["metadata"])
    path = os.path.join(OUTPUT_FOLDER, "metadata_output.pdf")
    pdf.output(path)
    return send_file(path, as_attachment=True)


@app.route("/api/reset", methods=["POST", "GET"])
@role_required(["Admin"])
def reset_metadata():
    global all_metadata
    all_metadata = []
    return jsonify({"message": "Server metadata reset successfully"})


@app.route("/", defaults={"path": ""})
@app.route("/<path:path>")
def serve_react(path):
    if path != "" and os.path.exists(os.path.join(app.static_folder, path)):
        return send_from_directory(app.static_folder, path)
    else:
        return send_from_directory(app.static_folder, "index.html")


if __name__ == "__main__":
    # Debug mode is fine for local development. In production, set debug=False.
    app.run(host="0.0.0.0", port=5000, debug=True)
